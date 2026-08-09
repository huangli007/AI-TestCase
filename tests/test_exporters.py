"""导出器测试:xlsx / md / json 结构验证。"""

import json
from pathlib import Path

from openpyxl import load_workbook

from testcase_agent.exporters import export_all


class TestExporters:
    def test_export_all(self, tmp_path, sample_result):
        paths = export_all(sample_result, str(tmp_path))
        assert len(paths) == 3
        for p in paths:
            assert Path(p).exists()

    def test_xlsx_structure(self, tmp_path, sample_result):
        export_all(sample_result, str(tmp_path), formats=["xlsx"])
        wb = load_workbook(next(tmp_path.glob("*.xlsx")))
        assert wb.sheetnames == ["产品分析", "测试点清单", "测试用例", "评审报告"]
        ws = wb["测试用例"]
        assert ws.max_row - 1 == len(sample_result.cases)  # 首行表头
        headers = [c.value for c in ws[1]]
        assert headers[0] == "用例编号" and headers[-1] == "预期结果"
        # 优先级着色存在
        cell = ws.cell(row=2, column=5)
        assert cell.value in ("P0", "P1", "P2", "P3")

    def test_markdown_content(self, tmp_path, sample_result):
        export_all(sample_result, str(tmp_path), formats=["md"])
        md = next(tmp_path.glob("*.md")).read_text(encoding="utf-8")
        assert "测试产品" in md
        assert "## 测试用例" in md
        assert "TC-001" in md and "正确验证码登录成功" in md

    def test_json_structure(self, tmp_path, sample_result):
        export_all(sample_result, str(tmp_path), formats=["json"])
        data = json.loads(next(tmp_path.glob("*.json")).read_text(encoding="utf-8"))
        assert data["product_name"] == "测试产品"
        assert len(data["cases"]) == len(sample_result.cases)
        assert "analysis" in data and "review" in data
        assert data["review"]["summary"] == "评审通过"

    def test_safe_name(self):
        from testcase_agent.exporters import safe_name
        assert safe_name('a/b\\c:d*e?f"g<h>i|j') == "a_b_c_d_e_f_g_h_i_j"
        assert safe_name("") == "product"
