"""Excel 导出:三张工作表(产品分析 / 测试点清单 / 测试用例),带样式。"""

from __future__ import annotations

import os
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .. import models

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIORITY_COLORS = {"P0": "FF6B6B", "P1": "FFA94D", "P2": "FFD43B", "P3": "69DB7C"}


def _style_sheet(ws, widths: List[int], n_rows: int, n_cols: int):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=n_cols):
        for cell in row:
            cell.border = BORDER
            cell.alignment = WRAP
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            else:
                cell.font = BODY_FONT
    ws.freeze_panes = "A2"


def _write_header(ws, headers: List[str]):
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)


def export_excel(result: models.GenerationResult, output_dir: str) -> str:
    wb = Workbook()

    # ---- Sheet 1: 产品分析 ----
    ws1 = wb.active
    ws1.title = "产品分析"
    analysis = result.analysis
    info_rows = [
        ("产品名称", analysis.product_name),
        ("产品类型", analysis.product_type),
        ("目标用户", analysis.target_users),
        ("核心价值", analysis.core_value),
        ("风险点", "\n".join(f"- {r}" for r in analysis.risk_points)),
        ("业务流", "\n".join(f"- {b.name}: {b.description}" for b in analysis.business_flows)),
    ]
    ws1.append(["项目", "内容"])
    for k, v in info_rows:
        ws1.append([k, v])
    ws1.append(["功能模块", ""])
    for m in analysis.modules:
        rules = "\n".join(f"  - {r}" for r in m.rules)
        ws1.append([f"{m.module} / {m.feature}", rules])
    _style_sheet(ws1, [24, 90], ws1.max_row, 2)

    # ---- Sheet 2: 测试点清单 ----
    ws2 = wb.create_sheet("测试点清单")
    _write_header(ws2, ["编号", "模块", "测试点", "类型", "优先级", "描述", "推断"])
    for tp in analysis.test_points:
        ws2.append([tp.id, tp.module, tp.name, tp.type, tp.priority, tp.description, "是" if tp.inferred else ""])
    _style_sheet(ws2, [10, 16, 22, 10, 9, 60, 8], ws2.max_row, 7)
    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value in PRIORITY_COLORS:
                cell.fill = PatternFill("solid", fgColor=PRIORITY_COLORS[cell.value])
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="333333")

    # ---- Sheet 3: 测试用例(核心) ----
    ws3 = wb.create_sheet("测试用例")
    _write_header(ws3, ["用例编号", "模块", "测试点", "用例标题", "优先级", "用例类型",
                        "前置条件", "测试数据", "操作步骤", "预期结果"])
    for c in result.cases:
        ws3.append([
            c.case_id, c.module, c.test_point, c.title, c.priority, c.case_type,
            c.precondition, c.test_data,
            "\n".join(f"{i}. {s}" for i, s in enumerate(c.steps, 1)),
            "\n".join(f"{i}. {e}" for i, e in enumerate(c.expected, 1)),
        ])
    _style_sheet(ws3, [10, 14, 18, 30, 9, 10, 24, 24, 48, 48], ws3.max_row, 10)
    for row in ws3.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value in PRIORITY_COLORS:
                cell.fill = PatternFill("solid", fgColor=PRIORITY_COLORS[cell.value])
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="333333")

    # ---- Sheet 4: 评审报告(如有) ----
    if result.review:
        ws4 = wb.create_sheet("评审报告")
        _write_header(ws4, ["项目", "内容"])
        ws4.append(["评审总结", result.review.summary])
        ws4.append(["补充的遗漏点", "\n".join(f"- {g}" for g in result.review.gaps)])
        ws4.append(["修正的问题", "\n".join(f"- {i}" for i in result.review.issues)])
        _style_sheet(ws4, [16, 100], ws4.max_row, 2)

    path = os.path.join(output_dir, f"{safe_name(result.product_name)}_测试用例.xlsx")
    wb.save(path)
    return path


def safe_name(name: str) -> str:
    return "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in name) or "product"
